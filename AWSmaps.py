import json
from openpyxl import load_workbook
from classes import Product
from mapping import PRODUCT_CATEGORY, PRODUCT_SERVICE, PRODUCT_TITLE, PRODUCT_RANKING

workbook = load_workbook(filename="C:/Users/lebon/AWSCloud/Data/class.xlsx", read_only=True)
sheet = workbook.active

products = []

# Using the values_only because you just want to return the cell value
for row in sheet.iter_rows(min_row=2, values_only=True):
    product = Product(Category=row[PRODUCT_CATEGORY],
                      Service=row[PRODUCT_SERVICE],
                      Title=row[PRODUCT_TITLE],
                      Ranking=row[PRODUCT_RANKING])
    products.append(product)

print()
print(products[0])
print()
print("Category: ",getattr(products[0],'Category'))
print("Service: ",getattr(products[0],'Service'))
print("Title: ",getattr(products[0],'Title'))
print("Ranking: ",getattr(products[0],'Ranking'))